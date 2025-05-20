# coding=gbk
import openpyxl
import json
import os
import shutil

def mode1():
    # ��ȡ Excel �ļ�
    wb = openpyxl.load_workbook("CLPD.xlsx")
    sheet = wb.active

    # ��ȡ�û�����
    train_rows = int(input("������ѵ����������"))
    val_rows = int(input("��������֤��������"))
    test_rows = int(input("��������Լ�������"))

    # ����Ŀ¼�ṹ
    base_dir = "D:/datasets/labels"
    dirs = ['train', 'val', 'test']
    [os.makedirs(os.path.join(base_dir, d), exist_ok=True) for d in dirs]

    current_row = 2  # �ӵڶ��п�ʼ
    for dir_name, max_rows in zip(dirs, [train_rows, val_rows, test_rows]):
        end_row = current_row + max_rows
        for idx, row in enumerate(sheet.iter_rows(min_row=current_row, max_row=end_row-1, values_only=True)):
            path, x1, y1, x2, y2, x3, y3, x4, y4, label = row

            json_data = {
                "shapes": [
                    {
                        "label": "single",
                        "points": [[x1, y1], [x2, y2], [x3, y3], [x4, y4]]
                    }
                ]
            }

            save_path = os.path.join(base_dir, dir_name)
            filename = os.path.join(save_path, f"{current_row-2}.json")
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(json_data, f)
            current_row += 1

def mode2():
    wb = openpyxl.load_workbook("CLPD.xlsx")
    sheet = wb.active

    # ��ȡ���б�ǩ
    labels = [row[9] for row in sheet.iter_rows(min_row=2, values_only=True)]

    # ����rec_trainĿ¼
    src_dir = "D:/datasets/images/rec_val"
    if not os.path.exists(src_dir):
        print(f"����Ŀ¼ {src_dir} ������")
        return

    files = sorted(os.listdir(src_dir), key=lambda x: int(x.split('.')[0]))
    for idx, filename in enumerate(files):
        if idx >= len(labels):
            break

        src = os.path.join(src_dir, filename)
        ext = os.path.splitext(filename)[1]
        new_name = f"{labels[idx]}_{idx}{ext}"
        dst = os.path.join(src_dir, new_name)
        os.rename(src, dst)
        print(f"������: {filename} -> {new_name}")

if __name__ == '__main__':
    choice = input("��ѡ��ģʽ (1-�����ݼ�����JSON / 2-������ͼƬ): ")

    if choice == '1':
        mode1()
    elif choice == '2':
        mode2()
    else:
        print("��Ч��ѡ��")

    print("������ɣ�")

