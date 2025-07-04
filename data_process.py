import openpyxl
import json
import os
from json2yolo import json2yolo

excel_dir =  r"D:/datasets/CLPD.xlsx"
img_dir = r"/datasets/images"
label_dir = r"D:/datasets/labels"
det_dir = r"D:/datasets/images/det_result"

_dir = ['train', 'val']
_dirs = ['train', 'val', 'test']
[os.makedirs(os.path.join(label_dir, d), exist_ok=True) for d in _dirs]



def get_json():
    wb = openpyxl.load_workbook(excel_dir)
    sheet = wb.active
    train_rows = int(input("请输入训练集图片数量："))
    val_rows = int(input("请输入验证集图片数量："))
    test_rows = int(input("请输入测试集图片数量："))
    current_row = 2  # 从第二行开始
    for dir_name, max_rows in zip(_dirs, [train_rows, val_rows, test_rows]):
        end_row = current_row + max_rows
        for idx, row in enumerate(sheet.iter_rows(min_row=current_row, max_row=end_row - 1, values_only=True)):
            path, x1, y1, x2, y2, x3, y3, x4, y4, label = row
            json_data = {
                "shapes": [
                    {
                        "label": "single",
                        "points": [[x1, y1], [x2, y2], [x3, y3], [x4, y4]]
                    }
                ]
            }
            save_path = os.path.join(label_dir, dir_name)
            filename = os.path.join(save_path, f"{current_row - 2}.json")
            with open(filename, "w", encoding="utf-8") as f:
                json.dump(json_data, f)
            current_row += 1


def get_name_label():
    wb = openpyxl.load_workbook(excel_dir)
    sheet = wb.active
    labels = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        filename = row[0].split('/')[1]
        if filename and row[9]:
            labels[filename] = row[9]
    return labels


def img_rename(det_dir):
    labels = get_name_label()
    for _, filename in enumerate(os.listdir(det_dir)):
        if labels[filename]:
            src = os.path.join(det_dir, filename)
            new_name = f"{labels[filename]}_{filename}"
            dst = os.path.join(det_dir, new_name)
            os.rename(src, dst)
            print(f"重命名: {filename} -> {new_name}")



if __name__ == '__main__':
    print('正在生成JSON文件并转换成Yolo格式')
    get_json()
    [json2yolo(os.path.join(img_dir, d), os.path.join(label_dir, d)) for d in _dir]
    print('操作完成')
